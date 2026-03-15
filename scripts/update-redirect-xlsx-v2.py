"""
Update Migration/Redirect .xlsx with tax page redirects.

Uses openpyxl to preserve all formatting, styles, and structure.
Creates a v2 copy per the document versioning rule.

Pass 1: Tax redirect updates (categories, items, gap analysis, cross-pillar, summary, FAQ rows)
Pass 2: Add human-readable name columns to FAQ Consolidation and Redirect Map sheets
"""

import re
import shutil
import os
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border

# ─── LC Category ID → Name lookup (from scraped folder names) ────────
LC_CAT_NAMES = {
    12: "General FAQs",
    13: "Business Entities Explained",
    14: "Business Taxation Explained",
    17: "General Incorporation FAQs",
    18: "Nevada Asset Protection Tutorial",
    19: "Why Incorporate in Nevada",
    21: "Limited Liability Company Tutorial",
    22: "Nevada Corporation Tutorial",
    27: "Nevada Asset Protection",
    29: "Nevada Asset Protection FAQs",
    41: "Company Taxation General FAQs",
    43: "LLC Taxation Tutorial",
    51: "Forms & Links Center",
    53: "St. Kitts and Nevis",
    54: "International Jurisdictions",
    55: "Corporation Taxation Tutorial",
    56: "Why Nevada",
    58: "Forms, Links & News Center",
    59: "Annual Services & Renewal FAQs",
    60: "Banking FAQs",
    61: "General Bank Account Opening FAQs",
    62: "International Asset Protection",
    63: "International Asset Protection Tutorial",
    64: "The Commonwealth of the Bahamas",
}

# Paths
BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SRC = os.path.join(
    BASE,
    "DOCS",
    "Phase 1 - Discovery & Site Architecture",
    "7 - Deliverables",
    "Incorporate123_Content_Migration_Map_and_Redirect_Plan.xlsx",
)
OUT = os.path.join(
    BASE,
    "DOCS",
    "Phase 1 - Discovery & Site Architecture",
    "7 - Deliverables",
    "v2-tax-redirects",
    "Incorporate123_Content_Migration_Map_and_Redirect_Plan_v2.xlsx",
)


def copy_cell_style(src_cell, dst_cell):
    """Copy all style attributes from src_cell to dst_cell."""
    dst_cell.font = copy(src_cell.font)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.alignment = copy(src_cell.alignment)
    dst_cell.border = copy(src_cell.border)
    dst_cell.number_format = src_cell.number_format


def copy_row_style(ws, src_row, dst_row):
    """Copy styling from every cell in src_row to dst_row."""
    for col in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(src_row, col), ws.cell(dst_row, col))


def apply_zebra(ws, row, is_even):
    """Apply zebra striping — even rows get light gray bg."""
    if is_even:
        gray = PatternFill("solid", fgColor="F5F5F5")
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = gray
    else:
        white = PatternFill("solid", fgColor="FFFFFF")
        for col in range(1, ws.max_column + 1):
            ws.cell(row, col).fill = white


# ─── Step 1: Copy source → v2, load workbook ────────────────────────
os.makedirs(os.path.dirname(OUT), exist_ok=True)
shutil.copy2(SRC, OUT)
wb = load_workbook(OUT)

# ─── Step 2: Update "Redirect Map" sheet ─────────────────────────────
ws = wb["Redirect Map"]

# 2a) Update 5 existing category rows
CATEGORY_UPDATES = {
    "/learning-center/cat/14": "/formation/entity-tax-guide/",
    "/learning-center/cat/59": "/compliance/tax-obligations/",
    "/learning-center/cat/41": "/compliance/tax-obligations/",
    "/learning-center/cat/43": "/formation/entity-tax-guide/",
    "/learning-center/cat/55": "/compare/llc-vs-scorp-vs-ccorp-tax/",
}

for row in range(5, ws.max_row + 1):
    old_url = ws.cell(row, 1).value
    if old_url in CATEGORY_UPDATES:
        ws.cell(row, 2).value = CATEGORY_UPDATES[old_url]

# 2b) Find the pattern rules row (first row with "Pattern" in col D)
pattern_row = None
for row in range(5, ws.max_row + 1):
    if ws.cell(row, 4).value == "Pattern":
        pattern_row = row
        break

assert pattern_row is not None, "Could not find pattern rules row"
print(f"Pattern rules start at row {pattern_row}")

# Define the 32 new item-level rows
NEW_REDIRECTS = [
    # To /formation/entity-tax-guide/ (13 rows)
    ("/learning-center/cat/14/item/37", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/14/item/40", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/14/item/62", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/14/item/88", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/14/item/89", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/43/item/37", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/43/item/40", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/43/item/62", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/43/item/88", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/43/item/89", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/21/item/37", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/41/item/88", "/formation/entity-tax-guide/"),
    ("/learning-center/cat/55/item/88", "/formation/entity-tax-guide/"),
    # To /compare/llc-vs-scorp-vs-ccorp-tax/ (9 rows)
    ("/learning-center/cat/14/item/39", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/14/item/90", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/14/item/91", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/43/item/39", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/43/item/90", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/43/item/91", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/55/item/39", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/55/item/91", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    ("/learning-center/cat/55/item/92", "/compare/llc-vs-scorp-vs-ccorp-tax/"),
    # To /compliance/tax-obligations/ (10 rows)
    ("/learning-center/cat/41/item/20", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/21", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/41", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/42", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/44", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/97", "/compliance/tax-obligations/"),
    ("/learning-center/cat/41/item/98", "/compliance/tax-obligations/"),
    ("/learning-center/cat/19/item/21", "/compliance/tax-obligations/"),
    ("/learning-center/cat/56/item/21", "/compliance/tax-obligations/"),
    ("/learning-center/cat/59/item/97", "/compliance/tax-obligations/"),
]

# Insert 32 blank rows before pattern rules
ws.insert_rows(pattern_row, len(NEW_REDIRECTS))

# Get a reference data row for styling (row just before the inserted block)
style_ref_row = pattern_row - 1

# Populate the new rows
for i, (old_url, new_url) in enumerate(NEW_REDIRECTS):
    row = pattern_row + i
    # Copy style from a nearby data row
    copy_row_style(ws, style_ref_row, row)
    # Set values
    ws.cell(row, 1).value = old_url       # Old URL Path
    ws.cell(row, 2).value = new_url        # New URL
    ws.cell(row, 3).value = None           # Inlinks (unknown for item-level)
    ws.cell(row, 4).value = "Individual"   # Redirect Type
    ws.cell(row, 5).value = "Medium"       # Priority
    # Apply zebra striping based on absolute row number
    apply_zebra(ws, row, row % 2 == 0)

print(f"Redirect Map: inserted {len(NEW_REDIRECTS)} rows, now {ws.max_row} rows")

# ─── Step 3: Update "Gap Analysis" sheet ─────────────────────────────
ws = wb["Gap Analysis"]

# Find the TOTAL row
total_row = None
for row in range(1, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and isinstance(val, str) and val.startswith("TOTAL NEW PAGES"):
        total_row = row
        break

assert total_row is not None, "Could not find TOTAL row in Gap Analysis"
print(f"Gap Analysis TOTAL at row {total_row}")

# Insert 3 rows before TOTAL
ws.insert_rows(total_row, 3)

# Style ref: last data row before insertion
style_ref = total_row - 1

GAP_ROWS = [
    (
        "/formation/entity-tax-guide/",
        "Entity Tax Guide",
        "Cluster",
        "Formation",
        "P2 \u2014 Expansion",
        "New creation + LC consolidation",
        "Consolidates tax-related LC items (cat/14, cat/43) into entity tax education page",
    ),
    (
        "/compare/llc-vs-scorp-vs-ccorp-tax/",
        "LLC vs S-Corp vs C-Corp Tax",
        "Comparison",
        "Formation",
        "P2 \u2014 Expansion",
        "New creation + LC consolidation",
        "Tax comparison across entity types. Consolidates LC items on tax treatment differences",
    ),
    (
        "/compliance/tax-obligations/",
        "Tax Filing & Obligations",
        "Cluster",
        "Compliance",
        "P2 \u2014 Expansion",
        "New creation + LC consolidation",
        "Consolidates tax obligation LC items (cat/41, cat/59) into compliance cluster",
    ),
]

p2_yellow = PatternFill("solid", fgColor="FFF9C4")

for i, (url, name, ptype, pillar, priority, source, notes) in enumerate(GAP_ROWS):
    row = total_row + i
    copy_row_style(ws, style_ref, row)
    ws.cell(row, 1).value = url
    ws.cell(row, 1).font = Font(name="Arial", size=10, bold=True)
    ws.cell(row, 2).value = name
    ws.cell(row, 3).value = ptype
    ws.cell(row, 4).value = pillar
    ws.cell(row, 5).value = priority
    ws.cell(row, 5).fill = p2_yellow
    ws.cell(row, 6).value = source
    ws.cell(row, 7).value = notes

# Update TOTAL row (now shifted down by 3)
new_total_row = total_row + 3
ws.cell(new_total_row, 1).value = "TOTAL NEW PAGES: 36"

print(f"Gap Analysis: added 3 rows, TOTAL now at row {new_total_row}, {ws.max_row} rows total")

# ─── Step 4: Update "Cross-Pillar References" sheet ──────────────────
ws = wb["Cross-Pillar References"]

# Append 2 rows after existing data
next_row = ws.max_row + 1
style_ref = ws.max_row  # last data row

XPILLAR_ROWS = [
    (
        "Entity Tax Guide",
        "/formation/entity-tax-guide/",
        "Formation",
        "Compliance",
        "CTA linking from Compliance tax sections; Formation pillar cluster nav",
        "Tax education spans formation decisions and compliance obligations",
    ),
    (
        "Tax Filing & Obligations",
        "/compliance/tax-obligations/",
        "Compliance",
        "Formation",
        "CTA linking from Formation entity pages; Compliance pillar cluster nav",
        "Tax obligations referenced from formation entity type pages",
    ),
]

for i, (page, url, primary, also_ref, strategy, notes) in enumerate(XPILLAR_ROWS):
    row = next_row + i
    copy_row_style(ws, style_ref, row)
    ws.cell(row, 1).value = page
    ws.cell(row, 2).value = url
    ws.cell(row, 3).value = primary
    ws.cell(row, 4).value = also_ref
    ws.cell(row, 5).value = strategy
    ws.cell(row, 6).value = notes
    apply_zebra(ws, row, row % 2 == 0)

print(f"Cross-Pillar References: now {ws.max_row} rows")

# ─── Step 5: Update "Migration Summary" dashboard ────────────────────
ws = wb["Migration Summary"]

SUMMARY_UPDATES = {
    "New Architecture Pages": "58\u201374+",
    "New Pages Required": "26+",
    "Cross-Pillar References": "14+",
}

for row in range(1, ws.max_row + 1):
    val = ws.cell(row, 1).value
    if val and isinstance(val, str) and val in SUMMARY_UPDATES:
        ws.cell(row, 2).value = SUMMARY_UPDATES[val]
        print(f"  Migration Summary row {row}: {val} = {SUMMARY_UPDATES[val]}")

# ─── Step 6: Update "FAQ Consolidation" sheet ─────────────────────────
ws = wb["FAQ Consolidation"]

next_row = ws.max_row + 1
style_ref = ws.max_row

FAQ_ROWS = [
    (14, "\u2014", "\u2014", "Tax-related", "/formation/entity-tax-guide/",
     "LLC taxation items redirect to entity tax guide instead of generic /faq/"),
    (43, "\u2014", "\u2014", "Tax-related", "/formation/entity-tax-guide/",
     "LLC taxation comparison items redirect to entity tax guide"),
    (55, "\u2014", "\u2014", "Tax-related", "/compare/llc-vs-scorp-vs-ccorp-tax/",
     "Corporation taxation items redirect to LLC vs S-Corp vs C-Corp tax comparison"),
    (41, "\u2014", "\u2014", "Tax-related", "/compliance/tax-obligations/",
     "Tax obligations items redirect to compliance tax obligations page"),
    (59, "\u2014", "\u2014", "Tax-related", "/compliance/tax-obligations/",
     "Use tax items redirect to compliance tax obligations page"),
]

for i, (cat_id, items, inlinks, pillar, target, rationale) in enumerate(FAQ_ROWS):
    row = next_row + i
    copy_row_style(ws, style_ref, row)
    ws.cell(row, 1).value = cat_id
    ws.cell(row, 2).value = items
    ws.cell(row, 3).value = inlinks
    ws.cell(row, 4).value = pillar
    ws.cell(row, 5).value = target
    ws.cell(row, 6).value = rationale
    apply_zebra(ws, row, row % 2 == 0)

print(f"FAQ Consolidation: now {ws.max_row} rows")

# ═══════════════════════════════════════════════════════════════════════
# PASS 2: Add human-readable name columns
# ═══════════════════════════════════════════════════════════════════════

# ─── Build URL → Name lookup from Migration Map ──────────────────────
ws_mm = wb["Migration Map"]
url_to_name = {}
for row in range(2, ws_mm.max_row + 1):
    url = ws_mm.cell(row, 1).value   # col A = Current URL
    name = ws_mm.cell(row, 7).value  # col G = New Page Name
    if url and name:
        url_to_name[url] = name


def get_redirect_page_name(old_url):
    """Derive a human-readable page name from an Old URL Path."""
    if not old_url or not isinstance(old_url, str):
        return None

    # LC item URLs: /learning-center/cat/XX/item/YY
    m = re.match(r"^/learning-center/cat/(\d+)/item/(\d+)$", old_url)
    if m:
        cat_id, item_id = int(m.group(1)), m.group(2)
        cat_name = LC_CAT_NAMES.get(cat_id, f"Category {cat_id}")
        return f"{cat_name} \u2014 Item {item_id}"

    # LC category URLs: /learning-center/cat/XX
    m = re.match(r"^/learning-center/cat/(\d+)$", old_url)
    if m:
        cat_id = int(m.group(1))
        return LC_CAT_NAMES.get(cat_id, f"Category {cat_id}")

    # SB60 legacy: /learning-center/sb60_2013_changes/cat/XX
    m = re.match(r"^/learning-center/sb60_2013_changes/cat/(\d+)$", old_url)
    if m:
        cat_id = int(m.group(1))
        return f"SB60 Legacy \u2014 cat/{cat_id}"

    # Pattern rows
    if old_url == "/learning-center/cat/*":
        return "All LC Categories (pattern)"
    if old_url == "/learning-center/cat/*/item/*":
        return "All LC Items (pattern)"
    if old_url == "/learning-center/sb60_2013_changes/*":
        return "All SB60 Pages (pattern)"
    if old_url == "/order?*":
        return "Order Query Strings (pattern)"

    # LC special pages (nrs-*, or /learning-center itself)
    if old_url == "/learning-center":
        return "Learning Center"
    m = re.match(r"^/learning-center/(.+)$", old_url)
    if m:
        slug = m.group(1)
        return slug.replace("-", " ").replace("_", " ").title()

    # Non-LC: look up in Migration Map
    name = url_to_name.get(old_url)
    if name:
        return name

    # Fallback: clean the URL slug
    slug = old_url.rstrip("/").rsplit("/", 1)[-1]
    return slug.replace("-", " ").replace("_", " ").title()


# ─── Step 8: FAQ Consolidation — Add "Category Name" column B ────────
ws = wb["FAQ Consolidation"]
faq_max_row = ws.max_row
faq_section2_start = 14  # CATEGORY-TO-PILLAR MAPPING header

# Handle merged cell at row 14 (A14:F14) — unmerge, will re-merge as A14:G14
if "A14:F14" in [str(m) for m in ws.merged_cells.ranges]:
    ws.unmerge_cells("A14:F14")

# Manually shift cells rightward in rows 14+ (don't use insert_cols which affects all rows)
# Shift right-to-left: G←F, F←E, E←D, D←C, C←B to make room for new col B
# Skip merged cells by checking cell type
from openpyxl.cell.cell import MergedCell
for row in range(faq_max_row, faq_section2_start - 1, -1):
    for col in range(7, 1, -1):  # 7←6, 6←5, 5←4, 4←3, 3←2
        src = ws.cell(row, col - 1)
        dst = ws.cell(row, col)
        if isinstance(src, MergedCell) or isinstance(dst, MergedCell):
            continue
        dst.value = src.value
        copy_cell_style(src, dst)
    # Clear col B (will be populated below)
    cell_b = ws.cell(row, 2)
    if not isinstance(cell_b, MergedCell):
        cell_b.value = None

# Re-merge row 14 with extended range (A14:G14)
ws.merge_cells("A14:G14")

# Set header
ws.cell(faq_section2_start + 1, 2).value = "Category Name"  # row 15
copy_cell_style(ws.cell(faq_section2_start + 1, 1), ws.cell(faq_section2_start + 1, 2))

# Populate Category Name for data rows (row 16 onward)
for row in range(faq_section2_start + 2, faq_max_row + 1):
    cat_id = ws.cell(row, 1).value
    if cat_id is not None:
        try:
            cat_id_int = int(cat_id)
            ws.cell(row, 2).value = LC_CAT_NAMES.get(cat_id_int, f"Category {cat_id_int}")
        except (ValueError, TypeError):
            pass
    # Style: copy from adjacent cell
    copy_cell_style(ws.cell(row, 3), ws.cell(row, 2))

print(f"FAQ Consolidation: added 'Category Name' column B (rows {faq_section2_start}–{faq_max_row})")

# ─── Step 9: Redirect Map — Add "Page Name" column B ─────────────────
ws = wb["Redirect Map"]

# Use insert_cols to add column B — this shifts all existing B→C, C→D, etc.
ws.insert_cols(2)

# Set header (row 4)
ws.cell(4, 2).value = "Page Name"
copy_cell_style(ws.cell(4, 1), ws.cell(4, 2))

# Populate Page Name for data rows (row 5 onward)
data_style = Font(name="Arial", size=10)
data_align = Alignment(vertical="top", wrap_text=True)

for row in range(5, ws.max_row + 1):
    old_url = ws.cell(row, 1).value
    name = get_redirect_page_name(old_url)
    ws.cell(row, 2).value = name
    ws.cell(row, 2).font = copy(data_style)
    ws.cell(row, 2).alignment = copy(data_align)
    # Preserve existing fill (zebra striping)
    ws.cell(row, 2).fill = copy(ws.cell(row, 1).fill)
    ws.cell(row, 2).border = copy(ws.cell(row, 1).border)

# Set column width for readability
ws.column_dimensions["B"].width = 35

print(f"Redirect Map: added 'Page Name' column B")

# ─── Step 10: Save ───────────────────────────────────────────────────
wb.save(OUT)
print(f"\nSaved: {OUT}")
